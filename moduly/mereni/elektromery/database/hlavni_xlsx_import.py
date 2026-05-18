from __future__ import annotations

import datetime
import io
import math
import re
from dataclasses import dataclass
from pathlib import Path
from collections.abc import Iterable
from typing import BinaryIO

import openpyxl
from sqlalchemy import inspect, text

from core.db.connect import ENGINE_MS, ENGINE_PG, SessionLocalMS, SessionLocalPG
from moduly.mereni.elektromery.database.models import (
    Elektromer_OTE_Mereni,
    Elektromer_areal_Zarizeni,
)
from moduly.mereni.elektromery.database.time_semantics import build_time_columns


HEADER_PATTERN = re.compile(
    r"seriove_cislo\s*:\s*(?P<serial>[^,]+)\s*,\s*identifikace\s*:\s*(?P<identifikace>.+)",
    re.IGNORECASE,
)
ARCHIVE_DIR = Path(__file__).resolve().parents[1] / "data"
INVALID_ARCHIVE_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
IDENTIFIKACE_CHECK_CHUNK_SIZE = 500


@dataclass(frozen=True)
class MainMeterDevice:
    identifikace: str
    seriove_cislo: int | None
    column_index: int
    source_label: str


@dataclass(frozen=True)
class MainMeterMeasurement:
    row_number: int
    identifikace: str
    seriove_cislo: int | None
    date: datetime.datetime
    objem: float


@dataclass(frozen=True)
class MainMeterImportIssue:
    message: str
    row_number: int | None = None
    identifikace: str | None = None
    date: datetime.datetime | None = None


@dataclass(frozen=True)
class ParsedMainMeterWorkbook:
    sheet_name: str
    devices: tuple[MainMeterDevice, ...]
    measurements: tuple[MainMeterMeasurement, ...]
    errors: tuple[MainMeterImportIssue, ...]
    warnings: tuple[MainMeterImportIssue, ...]


@dataclass(frozen=True)
class MainMeterImportResult:
    parsed: ParsedMainMeterWorkbook
    inserted_measurements: int
    skipped_existing_measurements: int
    conflict_measurements: int
    created_table: bool
    archived_file_path: str | None
    errors: tuple[MainMeterImportIssue, ...]
    warnings: tuple[MainMeterImportIssue, ...]


def _read_source_bytes(source: bytes | bytearray | BinaryIO) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, bytearray):
        return bytes(source)

    position = None
    if hasattr(source, "tell") and hasattr(source, "seek"):
        position = source.tell()
        source.seek(0)
    data = source.read()
    if position is not None:
        source.seek(position)
    return bytes(data)


def _parse_serial(value: object) -> int | None:
    if value is None:
        return None
    cleaned = str(value).strip().replace(" ", "")
    if not cleaned:
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _parse_datetime(value: object) -> datetime.datetime | None:
    if isinstance(value, datetime.datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)
    if value is None:
        return None

    text_value = str(value).strip()
    if not text_value:
        return None
    for date_format in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y",
    ):
        try:
            return datetime.datetime.strptime(text_value, date_format)
        except ValueError:
            continue
    return None


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        numeric_value = float(value)
        return None if math.isnan(numeric_value) else numeric_value

    text_value = str(value).strip().replace("\xa0", "").replace(" ", "")
    if not text_value:
        return None
    text_value = text_value.replace(",", ".")
    try:
        numeric_value = float(text_value)
    except ValueError:
        return None
    return None if math.isnan(numeric_value) else numeric_value


def _chunked(items: tuple[str, ...], size: int = IDENTIFIKACE_CHECK_CHUNK_SIZE):
    for start in range(0, len(items), size):
        yield items[start:start + size]


def _parse_header(value: object, column_index: int) -> tuple[MainMeterDevice | None, MainMeterImportIssue | None]:
    header = "" if value is None else str(value).strip()
    if not header:
        return None, None

    match = HEADER_PATTERN.search(header)
    if not match:
        return None, MainMeterImportIssue(
            message="Hlavicka sloupce neodpovida formatu 'seriove_cislo: ..., identifikace: ...'.",
            row_number=1,
        )

    identifikace = match.group("identifikace").strip()
    seriove_cislo = _parse_serial(match.group("serial"))
    if not identifikace:
        return None, MainMeterImportIssue(message="V hlavicce sloupce chybi identifikace.", row_number=1)

    return (
        MainMeterDevice(
            identifikace=identifikace,
            seriove_cislo=seriove_cislo,
            column_index=column_index,
            source_label=header,
        ),
        None,
    )


def parse_main_meter_xlsx(
    source: bytes | bytearray | BinaryIO,
    *,
    sheet_name: str | None = None,
) -> ParsedMainMeterWorkbook:
    workbook_bytes = _read_source_bytes(source)
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

    errors: list[MainMeterImportIssue] = []
    warnings: list[MainMeterImportIssue] = []
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        errors.append(MainMeterImportIssue(message="Soubor neobsahuje hlavickovy radek."))
        return ParsedMainMeterWorkbook(worksheet.title, (), (), tuple(errors), ())

    devices: list[MainMeterDevice] = []
    for column_index, header_value in enumerate(header_row[1:], start=2):
        device, issue = _parse_header(header_value, column_index)
        if issue is not None:
            errors.append(issue)
        if device is not None:
            devices.append(device)

    duplicate_devices = {
        device.identifikace
        for device in devices
        if sum(1 for item in devices if item.identifikace == device.identifikace) > 1
    }
    for identifikace in sorted(duplicate_devices):
        errors.append(
            MainMeterImportIssue(
                message="Identifikace je v hlavicce souboru uvedena vicekrat.",
                row_number=1,
                identifikace=identifikace,
            )
        )

    if not devices:
        errors.append(MainMeterImportIssue(message="Soubor neobsahuje zadny merici sloupec."))
        return ParsedMainMeterWorkbook(worksheet.title, (), (), tuple(errors), tuple(warnings))

    measurements: list[MainMeterMeasurement] = []
    seen_keys: set[tuple[str, datetime.datetime]] = set()

    for row_number, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_date = row_values[0] if row_values else None
        raw_values = [row_values[device.column_index - 1] if len(row_values) >= device.column_index else None for device in devices]
        if raw_date is None and all(value is None for value in raw_values):
            continue

        measurement_date = _parse_datetime(raw_date)
        if measurement_date is None:
            errors.append(MainMeterImportIssue(message="Radek obsahuje neplatne datum.", row_number=row_number))
            continue

        for device, raw_value in zip(devices, raw_values):
            objem = _parse_float(raw_value)
            if objem is None:
                warnings.append(
                    MainMeterImportIssue(
                        message="Prazdna nebo neciselna hodnota byla preskocena.",
                        row_number=row_number,
                        identifikace=device.identifikace,
                        date=measurement_date,
                    )
                )
                continue
            if objem < 0:
                errors.append(
                    MainMeterImportIssue(
                        message="Hodnota objem nemuze byt zaporna.",
                        row_number=row_number,
                        identifikace=device.identifikace,
                        date=measurement_date,
                    )
                )
                continue

            key = (device.identifikace, measurement_date)
            if key in seen_keys:
                errors.append(
                    MainMeterImportIssue(
                        message="Duplicitni mereni v importovanem souboru.",
                        row_number=row_number,
                        identifikace=device.identifikace,
                        date=measurement_date,
                    )
                )
                continue
            seen_keys.add(key)
            measurements.append(
                MainMeterMeasurement(
                    row_number=row_number,
                    identifikace=device.identifikace,
                    seriove_cislo=device.seriove_cislo,
                    date=measurement_date,
                    objem=round(objem, 6),
                )
            )

    return ParsedMainMeterWorkbook(
        sheet_name=worksheet.title,
        devices=tuple(devices),
        measurements=tuple(measurements),
        errors=tuple(errors),
        warnings=tuple(warnings),
    )


def ensure_elektromery_softlink_nullable() -> bool:
    inspector = inspect(ENGINE_MS)
    columns = inspector.get_columns("Mereni_elektromery", schema="dbo")
    softlink_column = next((column for column in columns if column["name"] == "softlink_id"), None)
    if softlink_column is None or softlink_column.get("nullable"):
        return False

    with ENGINE_MS.begin() as connection:
        connection.execute(text("ALTER TABLE dbo.Mereni_elektromery ALTER COLUMN softlink_id BIGINT NULL"))
    return True


def ensure_elektromery_ote_table() -> bool:
    with ENGINE_PG.begin() as connection:
        connection.execute(text("CREATE SCHEMA IF NOT EXISTS dbo"))

    inspector = inspect(ENGINE_PG)
    table_name = Elektromer_OTE_Mereni.__tablename__
    if table_name in inspector.get_table_names(schema="dbo"):
        with ENGINE_PG.begin() as connection:
            connection.execute(
                text(
                    """
                    ALTER TABLE dbo."Mereni_elektromery_OTE"
                        ADD COLUMN IF NOT EXISTS source_date TIMESTAMP WITHOUT TIME ZONE,
                        ADD COLUMN IF NOT EXISTS time_utc TIMESTAMP WITH TIME ZONE,
                        ADD COLUMN IF NOT EXISTS time_basis VARCHAR(40),
                        ADD COLUMN IF NOT EXISTS source_timezone VARCHAR(64),
                        ADD COLUMN IF NOT EXISTS source_utc_offset_minutes INTEGER,
                        ADD COLUMN IF NOT EXISTS time_fold INTEGER,
                        ADD COLUMN IF NOT EXISTS timestamp_position VARCHAR(20)
                    """
                )
            )
            connection.execute(
                text(
                    """
                    CREATE INDEX IF NOT EXISTS ix_ele_ote_time_utc
                    ON dbo."Mereni_elektromery_OTE" (time_utc)
                    """
                )
            )
        return False

    Elektromer_OTE_Mereni.__table__.create(bind=ENGINE_PG, checkfirst=True)
    with ENGINE_PG.begin() as connection:
        connection.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS ix_ele_ote_time_utc
                ON dbo."Mereni_elektromery_OTE" (time_utc)
                """
            )
        )
    return True


def find_unknown_elektromery_identifikace(identifikace_values: Iterable[str]) -> tuple[str, ...]:
    requested_identifikace = tuple(
        sorted(
            {
                str(identifikace).strip()
                for identifikace in identifikace_values
                if str(identifikace).strip()
            }
        )
    )
    if not requested_identifikace:
        return ()

    existing_identifikace: set[str] = set()
    with SessionLocalMS() as session:
        for identifikace_chunk in _chunked(requested_identifikace):
            existing_identifikace.update(
                row[0]
                for row in session.query(Elektromer_areal_Zarizeni.identifikace)
                .filter(Elektromer_areal_Zarizeni.identifikace.in_(identifikace_chunk))
                .all()
            )

    return tuple(
        identifikace
        for identifikace in requested_identifikace
        if identifikace not in existing_identifikace
    )


def build_unknown_identification_issues(parsed: ParsedMainMeterWorkbook) -> tuple[MainMeterImportIssue, ...]:
    unknown_identifikace = find_unknown_elektromery_identifikace(
        device.identifikace
        for device in parsed.devices
    )
    return tuple(
        MainMeterImportIssue(
            message="Identifikace neni zalozena v MS tabulce dbo.Zarizeni_elektromery.",
            row_number=1,
            identifikace=identifikace,
        )
        for identifikace in unknown_identifikace
    )


def _safe_archive_filename(source_file: str | None) -> str:
    raw_name = Path(str(source_file or "elektromery_import.xlsx")).name.strip()
    if not raw_name:
        raw_name = "elektromery_import.xlsx"
    safe_name = INVALID_ARCHIVE_FILENAME_CHARS.sub("_", raw_name).strip(" .")
    if not safe_name:
        safe_name = "elektromery_import.xlsx"
    if not safe_name.lower().endswith(".xlsx"):
        safe_name = f"{safe_name}.xlsx"
    return safe_name


def archive_main_meter_xlsx(
    source: bytes | bytearray | BinaryIO,
    *,
    source_file: str | None = None,
    archive_dir: str | Path | None = None,
) -> Path:
    workbook_bytes = _read_source_bytes(source)
    resolved_archive_dir = Path(archive_dir) if archive_dir is not None else ARCHIVE_DIR
    resolved_archive_dir.mkdir(parents=True, exist_ok=True)

    safe_name = _safe_archive_filename(source_file)
    target_path = resolved_archive_dir / safe_name
    if target_path.exists():
        if target_path.read_bytes() == workbook_bytes:
            return target_path

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = target_path.stem
        suffix = target_path.suffix
        for index in range(1, 1000):
            counter = "" if index == 1 else f"_{index}"
            candidate_path = resolved_archive_dir / f"{stem}_{timestamp}{counter}{suffix}"
            if not candidate_path.exists():
                target_path = candidate_path
                break
        else:
            raise FileExistsError(f"Nepodarilo se najit volny nazev pro archivaci souboru {safe_name}.")

    target_path.write_bytes(workbook_bytes)
    return target_path


def _values_match(left: object, right: object) -> bool:
    left_value = _parse_float(left)
    right_value = _parse_float(right)
    if left_value is None or right_value is None:
        return left_value is right_value
    return abs(left_value - right_value) <= 0.000001


def import_main_meter_xlsx(
    source: bytes | bytearray | BinaryIO,
    *,
    sheet_name: str | None = None,
    source_file: str | None = None,
    archive_dir: str | Path | None = None,
) -> MainMeterImportResult:
    workbook_bytes = _read_source_bytes(source)
    parsed = parse_main_meter_xlsx(workbook_bytes, sheet_name=sheet_name)
    if parsed.errors:
        return MainMeterImportResult(
            parsed=parsed,
            inserted_measurements=0,
            skipped_existing_measurements=0,
            conflict_measurements=0,
            created_table=False,
            archived_file_path=None,
            errors=parsed.errors,
            warnings=parsed.warnings,
        )

    identification_errors = build_unknown_identification_issues(parsed)
    if identification_errors:
        return MainMeterImportResult(
            parsed=parsed,
            inserted_measurements=0,
            skipped_existing_measurements=0,
            conflict_measurements=0,
            created_table=False,
            archived_file_path=None,
            errors=identification_errors,
            warnings=parsed.warnings,
        )

    created_table = ensure_elektromery_ote_table()
    errors: list[MainMeterImportIssue] = []
    warnings = list(parsed.warnings)
    skipped_existing = 0
    conflict_count = 0
    inserted_count = 0

    with SessionLocalPG() as session:
        if not parsed.measurements:
            return MainMeterImportResult(
                parsed=parsed,
                inserted_measurements=0,
                skipped_existing_measurements=0,
                conflict_measurements=0,
                created_table=created_table,
                archived_file_path=None,
                errors=(),
                warnings=tuple(warnings),
            )

        identifiers = tuple(sorted({measurement.identifikace for measurement in parsed.measurements}))
        start_date = min(measurement.date for measurement in parsed.measurements)
        end_date = max(measurement.date for measurement in parsed.measurements)
        existing_rows = (
            session.query(
                Elektromer_OTE_Mereni.identifikace,
                Elektromer_OTE_Mereni.date,
                Elektromer_OTE_Mereni.objem,
            )
            .filter(
                Elektromer_OTE_Mereni.identifikace.in_(identifiers),
                Elektromer_OTE_Mereni.date >= start_date,
                Elektromer_OTE_Mereni.date <= end_date,
            )
            .all()
        )
        existing_by_key = {(row.identifikace, row.date): row.objem for row in existing_rows}

        new_rows: list[Elektromer_OTE_Mereni] = []
        for measurement in parsed.measurements:
            measurement_key = (measurement.identifikace, measurement.date)
            if measurement_key in existing_by_key:
                existing_objem = existing_by_key[measurement_key]
                if _values_match(existing_objem, measurement.objem):
                    skipped_existing += 1
                else:
                    conflict_count += 1
                    errors.append(
                        MainMeterImportIssue(
                            message="V databazi uz existuje mereni se stejnou identifikaci a datem, ale jinou hodnotou.",
                            row_number=measurement.row_number,
                            identifikace=measurement.identifikace,
                            date=measurement.date,
                        )
                    )
                continue

            new_rows.append(
                Elektromer_OTE_Mereni(
                    identifikace=measurement.identifikace,
                    seriove_cislo=measurement.seriove_cislo,
                    objem=measurement.objem,
                    date=measurement.date,
                    **build_time_columns(measurement.date, "OTE"),
                    source_file=source_file,
                )
            )

        if errors:
            session.rollback()
        else:
            session.add_all(new_rows)
            inserted_count = len(new_rows)
            session.commit()

    archived_file_path = None
    if not errors:
        try:
            archived_file_path = str(
                archive_main_meter_xlsx(
                    workbook_bytes,
                    source_file=source_file,
                    archive_dir=archive_dir,
                )
            )
        except OSError as exc:
            warnings.append(
                MainMeterImportIssue(
                    message=f"Import byl ulozen, ale soubor se nepodarilo archivovat: {exc}",
                )
            )

    return MainMeterImportResult(
        parsed=parsed,
        inserted_measurements=inserted_count,
        skipped_existing_measurements=skipped_existing,
        conflict_measurements=conflict_count,
        created_table=created_table,
        archived_file_path=archived_file_path,
        errors=tuple(errors),
        warnings=tuple(warnings),
    )
